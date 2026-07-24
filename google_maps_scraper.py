import time
import re
import logging
from dataclasses import dataclass
from typing import Optional

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    NoSuchElementException,
    TimeoutException,
    StaleElementReferenceException,
)
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


SEARCH_QUERY = "restaurants in New York"
MAX_RESULTS = 60
OUTPUT_FILE = "nyc_restaurants.xlsx"
SCROLL_PAUSE = 2.0
PAGE_LOAD_WAIT = 10
HEADLESS = False


logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)


@dataclass
class Restaurant:
    name: str = "N/A"
    address: str = "N/A"
    phone: str = "N/A"
    rating: Optional[float] = None
    review_count: int = 0
    website: str = "N/A"
    cuisine: str = "N/A"
    maps_url: str = "N/A"


def build_driver(headless: bool = False) -> webdriver.Chrome:
    """Return a Chrome driver configured to reduce bot detection."""
    options = Options()

    if headless:
        options.add_argument("--headless=new")

    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("--window-size=1400,900")
    options.add_argument("--lang=en-US")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option("useAutomationExtension", False)
    options.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/122.0.0.0 Safari/537.36"
    )

    driver = webdriver.Chrome(options=options)
    driver.execute_script(
        "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
    )
    log.info("Chrome WebDriver started.")
    return driver


def safe_find_text(driver, css_selector: str, default: str = "N/A") -> str:
    try:
        el = driver.find_element(By.CSS_SELECTOR, css_selector)
        return el.text.strip() or default
    except NoSuchElementException:
        return default


def parse_rating(raw: str) -> Optional[float]:
    try:
        return float(raw.replace(",", "."))
    except (ValueError, AttributeError):
        return None


def parse_review_count(raw: str) -> int:
    """Pull an integer out of strings like '(1,234)' or '1234 reviews'."""
    digits = re.sub(r"[^\d]", "", raw)
    return int(digits) if digits else 0


def scrape_google_maps(query: str, max_results: int) -> list[Restaurant]:
    """Search Google Maps, scroll the results panel, and extract each place."""
    driver = build_driver(headless=HEADLESS)
    wait = WebDriverWait(driver, PAGE_LOAD_WAIT)
    results: list[Restaurant] = []

    try:
        log.info("Opening Google Maps")
        driver.get("https://www.google.com/maps?hl=en")
        time.sleep(4)

        # Cookie consent appears in EU regions
        try:
            consent_btn = driver.find_element(
                By.XPATH, '//button[contains(., "Accept all") or contains(., "Reject all")]'
            )
            consent_btn.click()
            time.sleep(1)
        except NoSuchElementException:
            pass

        log.info(f"Searching for: '{query}'")

        # Google rotates these, so try them in order
        search_box = None
        selectors = [
            (By.ID, "searchboxinput"),
            (By.NAME, "q"),
            (By.CSS_SELECTOR, 'input[type="text"]'),
            (By.CSS_SELECTOR, 'input[aria-label]'),
        ]

        for by, selector in selectors:
            try:
                search_box = WebDriverWait(driver, 8).until(
                    EC.presence_of_element_located((by, selector))
                )
                log.info(f"Search box found via: {selector}")
                break
            except TimeoutException:
                log.warning(f"Selector not found: {selector}, trying next")

        if search_box is None:
            log.error("Could not find search box, saving screenshot for debug")
            driver.save_screenshot("debug_screenshot.png")
            log.error("Screenshot saved as debug_screenshot.png")
            return results

        search_box.clear()
        search_box.send_keys(query)
        time.sleep(1)
        search_box.send_keys(Keys.ENTER)
        time.sleep(4)

        log.info("Scrolling results panel to load listings")
        results_panel_selector = 'div[role="feed"]'

        try:
            panel = wait.until(
                EC.presence_of_element_located((By.CSS_SELECTOR, results_panel_selector))
            )
        except TimeoutException:
            log.warning("Results panel not found, the search may have returned nothing")
            return results

        last_count = 0
        stall_counter = 0

        while len(results) < max_results and stall_counter < 5:
            cards = driver.find_elements(
                By.CSS_SELECTOR, 'a[href*="/maps/place/"]'
            )
            current_count = len(cards)

            log.info(f"  Loaded {current_count} listing cards")

            driver.execute_script("arguments[0].scrollTop += 1200;", panel)
            time.sleep(SCROLL_PAUSE)

            # No new cards for 5 rounds means the feed is exhausted
            if current_count == last_count:
                stall_counter += 1
            else:
                stall_counter = 0
                last_count = current_count

        # Collect hrefs as plain strings first, then visit each URL directly.
        # Holding element references across navigation is what causes
        # StaleElementReferenceException; strings have no DOM dependency.
        log.info("Collecting place URLs from results panel")

        place_urls = []
        seen = set()

        all_anchors = driver.find_elements(By.CSS_SELECTOR, 'a[href*="/maps/place/"]')
        for a in all_anchors:
            try:
                href = a.get_attribute("href")
                if href and href not in seen:
                    seen.add(href)
                    place_urls.append(href)
            except StaleElementReferenceException:
                continue

        place_urls = place_urls[:max_results]
        log.info(f"Collected {len(place_urls)} unique place URLs, starting extraction")

        for idx, place_url in enumerate(place_urls, start=1):
            restaurant = Restaurant()
            restaurant.maps_url = place_url

            try:
                driver.get(place_url)
                time.sleep(3)

                try:
                    name_el = wait.until(
                        EC.presence_of_element_located(
                            (By.CSS_SELECTOR, 'h1.DUwDvf, h1[class*="fontHeadlineLarge"]')
                        )
                    )
                    restaurant.name = name_el.text.strip()
                except TimeoutException:
                    restaurant.name = "N/A"

                try:
                    rating_el = driver.find_element(
                        By.CSS_SELECTOR,
                        'div.F7nice span[aria-hidden="true"]'
                    )
                    restaurant.rating = parse_rating(rating_el.text)
                except NoSuchElementException:
                    restaurant.rating = None

                try:
                    reviews_el = driver.find_element(
                        By.CSS_SELECTOR,
                        'div.F7nice span[aria-label*="review"]'
                    )
                    restaurant.review_count = parse_review_count(
                        reviews_el.get_attribute("aria-label")
                    )
                except NoSuchElementException:
                    restaurant.review_count = 0

                try:
                    cuisine_el = driver.find_element(
                        By.CSS_SELECTOR,
                        'button.DkEaL, span.mgr77e'
                    )
                    restaurant.cuisine = cuisine_el.text.strip()
                except NoSuchElementException:
                    restaurant.cuisine = "N/A"

                try:
                    addr_el = driver.find_element(
                        By.CSS_SELECTOR,
                        'button[data-item-id="address"] .Io6YTe'
                    )
                    restaurant.address = addr_el.text.strip()
                except NoSuchElementException:
                    restaurant.address = "N/A"

                try:
                    phone_el = driver.find_element(
                        By.CSS_SELECTOR,
                        'button[data-item-id*="phone"] .Io6YTe'
                    )
                    restaurant.phone = phone_el.text.strip()
                except NoSuchElementException:
                    restaurant.phone = "N/A"

                try:
                    web_el = driver.find_element(
                        By.CSS_SELECTOR,
                        'a[data-item-id="authority"] .Io6YTe'
                    )
                    restaurant.website = web_el.text.strip()
                except NoSuchElementException:
                    restaurant.website = "N/A"

                results.append(restaurant)
                log.info(
                    f"  [{idx}/{len(place_urls)}] {restaurant.name} "
                    f"| {restaurant.rating} ({restaurant.review_count} reviews)"
                )

            except (StaleElementReferenceException, TimeoutException) as exc:
                log.warning(f"  [{idx}] Skipped: {type(exc).__name__}")
                continue

    finally:
        driver.quit()
        log.info("Browser closed.")

    return results


def sort_results(data: list[Restaurant]) -> list[Restaurant]:
    """Sort by rating, then review count. Unrated places go last."""
    return sorted(
        data,
        key=lambda r: (r.rating is not None, r.rating or 0, r.review_count),
        reverse=True,
    )


def export_to_excel(data: list[Restaurant], filename: str) -> None:
    """Write results to a styled workbook with a frozen header and a summary row."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "NYC Restaurants"

    headers = [
        "#", "Name", "Cuisine", "Rating", "Reviews",
        "Address", "Phone", "Website", "Google Maps URL"
    ]

    HEADER_BG = "1A73E8"
    HEADER_FONT = "FFFFFF"
    ROW_ALT_BG = "EAF1FB"
    BORDER_COLOR = "BFCFE8"

    thin_border = Border(
        left=Side(style="thin", color=BORDER_COLOR),
        right=Side(style="thin", color=BORDER_COLOR),
        top=Side(style="thin", color=BORDER_COLOR),
        bottom=Side(style="thin", color=BORDER_COLOR),
    )

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color=HEADER_FONT, size=11)
        cell.fill = PatternFill("solid", fgColor=HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    ws.row_dimensions[1].height = 30

    for row_idx, r in enumerate(data, start=2):
        row_data = [
            row_idx - 1,
            r.name,
            r.cuisine,
            r.rating if r.rating is not None else "N/A",
            r.review_count,
            r.address,
            r.phone,
            r.website,
            r.maps_url,
        ]

        is_alt_row = (row_idx % 2 == 0)
        fill = PatternFill("solid", fgColor=ROW_ALT_BG) if is_alt_row else None

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = thin_border
            if fill:
                cell.fill = fill

            # Highlight strong ratings
            if col_idx == 4 and isinstance(value, float) and value >= 4.5:
                cell.font = Font(bold=True, color="0A6E0A")

        ws.row_dimensions[row_idx].height = 22

    col_widths = [5, 32, 20, 10, 12, 38, 16, 28, 50]
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"

    summary_row = len(data) + 2
    ws.cell(row=summary_row, column=1, value="Total").font = Font(bold=True)
    ws.cell(row=summary_row, column=2, value=len(data)).font = Font(bold=True)
    avg_rating = (
        sum(r.rating for r in data if r.rating) / max(1, sum(1 for r in data if r.rating))
    )
    ws.cell(row=summary_row, column=4, value=round(avg_rating, 2)).font = Font(bold=True)

    wb.save(filename)
    log.info(f"Excel report saved to {filename}")


def main():
    log.info("Google Maps Scraper")

    raw_data = scrape_google_maps(
        query=SEARCH_QUERY,
        max_results=MAX_RESULTS,
    )

    if not raw_data:
        log.error("No data collected. Check the connection or the selectors.")
        return

    log.info(f"Collected {len(raw_data)} restaurants, sorting")

    sorted_data = sort_results(raw_data)
    export_to_excel(sorted_data, OUTPUT_FILE)

    log.info(f"Done. Results are in {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
